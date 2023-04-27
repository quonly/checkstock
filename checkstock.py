import time
import random
from datetime import datetime
from openpyxl import load_workbook, Workbook
from colorama import  just_fix_windows_console,init
from colorama import Fore,Back, Style
'''
pyinstaller --onefile checkstock.py ทำเป็นไฟล์ exe

Available formatting constants are:
Fore: BLACK, RED, GREEN, YELLOW, BLUE, MAGENTA, CYAN, WHITE, RESET.
Back: BLACK, RED, GREEN, YELLOW, BLUE, MAGENTA, CYAN, WHITE, RESET.
Style: DIM, NORMAL, BRIGHT, RESET_ALL
'''
just_fix_windows_console()
print(Style.BRIGHT)
file_name = input(f"ชื่อไฟล์ excel(xlsx):")
file = load_workbook(file_name+'.xlsx')
sheet = file.worksheets[0]
max_row = sheet.max_row
col_onhand = int(input(f'คอลัมน์ onhand:'))
col_outgoing = int(input('คอลัมน์ outgoing:'))
export = []
is_found = True
execute = True
while execute:
    if not is_found:
        print(f'{Fore.WHITE}{Back.RED}ไม่พบสินค้า{Style.RESET_ALL}')
    print(f'{Fore.RED}กด 0 จบการทำงาน{Style.RESET_ALL}\n{Fore.MAGENTA}กด 1 export ผลลัพธ์{Style.RESET_ALL}')
    search = input("SKU: ")
    # num_row จะเริ่มนับจาก 0

    match search:
        case '0':
            print('จบการทำงาน')
            break
        case '1':
            crr_d = datetime.now()
            day = crr_d.day
            month = crr_d.month
            year = crr_d.year
            yearMonthDay = f'{year}-{month}-{day}'
            excel = Workbook()
            ran = random.randrange(1,1000)
            excel_name = f"stock-export{yearMonthDay}{ran}.xlsx"
            excel.save(excel_name)
            work_sheet = excel.worksheets[0]
            work_sheet.title = 'สินค้าคงคลัง'
            header = ['SKU','ONHAND','OUTGOING','BALANCE']
            '''
      export เป็น list 2 มิติ 
      ข้อมูลที่ต้องการจริงๆ คือ ลิสต์ที่ 2
      จึงต้องนำจวนทั้งหมดที่มีออกมาก่อน
      แล้วจึงเข้าถึงข้อมูลภายในด้วยลำดับที่ได้จากจำนวนด้านบน
      len จะทำให้เราได้จำนวนของชุดข้อมูลที่อยู่ภายใน export
      len ได้ข้อมูลออกมาเป็น int ซึ่งไม่สามารถวนรอบได้
      range จะทำให้ได้ข้อมูล 0 - จำนวน range ที่ใส่เข้าไป range(5) -> [0,5]
      '''   
            r = 2
            for items in export:
                for i in range(4):
                    if (r-1) == 1:
                        work_sheet.cell(row=1, column=i+1).value = header[i]
                        work_sheet.cell(row=r, column=i+1).value = items[i]
                    else:
                        work_sheet.cell(row=r, column=i+1).value = items[i]
                r += 1
            excel.save(excel_name)
            print(f'{Back.GREEN}{Style.BRIGHT}{Fore.WHITE}EXPORT SUCCESS.{Style.RESET_ALL}')
            execute = False
        case _:
            for num_row in range(max_row):
                # จะวนรอบตาม max_row
                # ให้เข้าถึง cell ของแต่ละ row ในแต่ละรอบ
                # cell(row=row,column=2) method cell จะเริ่มอ่านจากแถว(row) แล้วจึงเป็นคอลัมน์(column) เช่น
                # A1 row=1 , column=1
                # A2 row=1 , column=2
                # B1 row=2 , column=1
                row = num_row + 1
                # print(row)
                # เริ่มต้นนับ row จาก 1
                sku = str(sheet.cell(row=row, column=2).value).lower()
                search_sku = search.lower()
                # print(sku)
                if search_sku in sku:
                    is_found = True
                    onhand = int(sheet.cell(
                        row=row, column=col_onhand).value) or 0
                    outgoing = int(sheet.cell(
                        row=row, column=col_outgoing).value) or 0
                    result = onhand-outgoing
                    print(f'{Fore.WHITE}{Back.GREEN}{Style.BRIGHT}สินค้าคงเหลือ: {result}{Style.RESET_ALL}')
                    items = [search_sku.upper(), onhand, outgoing, result]
                    export.append(items)
                    check = 1
                    break
                is_found = False
