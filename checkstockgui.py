import tkinter as tk
from datetime import datetime
from tkinter import filedialog,ttk
import random
from openpyxl import load_workbook,Workbook


class File:
    export_items = [['SKU','ONHAND','OUTGOING','BALANCE']]
    file_path = None
    max_row = 0
    sheet = None
    col_sku = None
    col_onhand = None
    col_out_going = None
    crr_d = datetime.now()
    ran = random.randrange(1,100000)
    
    day = crr_d.day
    month = crr_d.month
    year = crr_d.year
    yearMonthDay = f'{year}-{month}-{day}'
    
    def filePath(self):
        self.file_path = filedialog.askopenfilename()
        result_openfile.configure(text=f'เปิดไฟล์: {self.file_path}')
        self.openFile()

    def openFile(self):
        file = load_workbook(self.file_path)
        self.sheet = file.worksheets[0]
        self.max_row = self.sheet.max_row
    
    def setCol(self,col_sku,col_onhand,col_out_going):
        try:
            self.col_onhand = int(col_onhand)
            self.col_out_going = int(col_out_going)
            self.col_sku = int(col_sku)
        except ValueError:
            result_label.configure(text=f'คอลัมน์ไม่ถูกต้อง')
        
    def exportFile(self,filename=f"stock-export{yearMonthDay}{ran}.xlsx",r=1):
        if len(self.export_items) > 1:
            excel = Workbook()
            excel.save(filename)
            work_sheet = excel.worksheets[0]
            r=1
            for item in self.export_items:
                for i in range(4):
                    work_sheet.cell(row=r, column=i+1).value = item[i]
                r += 1
            excel.save(filename)  
            result_label.configure(text=f'Export สำเร็จ: {filename}')
        else:result_label.configure(text=f'ไม่พบข้อมูลที่สามารถ export ได้')
class CheckStock(File):
    def search_kw(self,kw):
        self.kw = kw  
        return self.kw
    def stock(self,event=None):
        self.setCol(col_sku_entry.get(),onhand_entry.get(),outgoing_entry.get())
        search = self.search_kw(sku_entry.get())
        s_sku = search.lower().replace('#','')
        is_found = False
        for row in range(self.max_row):
            row+=1
            sku = str(self.sheet.cell(row=row, column=self.col_sku).value).lower().replace('#','')
            if s_sku == sku:
                onhand = int(self.sheet.cell(row=row, column=self.col_onhand).value) or 0
                outgoing = int(self.sheet.cell(row=row, column=self.col_out_going).value) or 0
                result = onhand - outgoing
                result_label.configure(text=f'{s_sku.upper()} สินค้าคงเหลือ: {result}')
                items = [s_sku.upper(), onhand, outgoing, result]
                if items not in self.export_items:
                    self.export_items.append(items)
                is_found = True
                break
        if not is_found:
            result_label.configure(text=f'ไม่พบสินค้า: {s_sku}')    

root = tk.Tk()
root.geometry('600x800')
root.title('Search SKU')
root.iconbitmap('icon/favicon.ico')
root.option_add('*Font','Angsana_New 14')


howto_text = tk.StringVar()
howto_text.set('วิธีการใช้\n1. เปิดไฟล์ที่ต้องการด้วยปุ่ม Open file\n2. คอลัมน์ onhand ใส่เลขคอลัมน์ของปริมาณในมือ\n3. คอลัมน์ outgoing ใส่เลขคอลัมน์ของกล่องขาออก\n4. กรอกรหัสสินค้าที่ช่อง SKU กด Enter\n5. Export file จะส่งออกไฟล์ excel ตามข้อมูลสินค้าที่ค้นพบ')
howto = tk.Label(root,textvariable=howto_text,justify='left',pady=20,height=8)
howto.pack()



button_frame = ttk.Frame(root)
button_frame.pack(pady=10)

f = CheckStock()
openfile_button = ttk.Button(button_frame, text='Open file', command=f.filePath)
openfile_button.pack(side='left')

export_button = ttk.Button(button_frame, text='Export file', command=f.exportFile)
export_button.pack(side='left',padx=10)

result_openfile = tk.Label(root, text='', width=50, height=2)
result_openfile.pack(pady=5,ipady=10)

col_sku_label = tk.Label(root, text='คอลัมน์ SKU:')
col_sku_label.pack(pady=5)
col_sku_entry = ttk.Entry(root)
col_sku_entry.pack()

onhand_label = tk.Label(root, text='คอลัมน์ onhand:')
onhand_label.pack(pady=5)

onhand_entry = ttk.Entry(root)
onhand_entry.pack()

outgoing_label = ttk.Label(root, text='คอลัมน์ outgoing:')
outgoing_label.pack(pady=5)

outgoing_entry = ttk.Entry(root)
outgoing_entry.pack()

sku_label = tk.Label(root, text='SKU:')
sku_label.pack(pady=5)

sku_entry = ttk.Entry(root)
sku_entry.pack()

search_button = ttk.Button(root, text='Search', command=f.stock)
search_button.pack(pady=10)
sku_entry.bind('<Return>',f.stock)

result_label = tk.Label(root, text='', width=50, height=2)
result_label.pack(pady=20,ipady=10)

root.mainloop()
