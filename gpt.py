import time
import random
from datetime import datetime
from openpyxl import load_workbook, Workbook
from colorama import  just_fix_windows_console,init
from colorama import Fore,Back, Style

just_fix_windows_console()
print(Style.BRIGHT)
file_name = input(f"ชื่อไฟล์ excel(xlsx):")
file = load_workbook(file_name+'.xlsx')
sheet = file.worksheets[0]
max_row = sheet.max_row
col_onhand = int(input(f'คอลัมน์ onhand:'))
col_outgoing = int(input('คอลัมน์ outgoing:'))
export = []
is_found = True
execute = True
# change while loop to funtion inside button (submit)
while execute:
    if not is_found:
        print(f'{Fore.WHITE}{Back.RED}ไม่พบสินค้า{Style.RESET_ALL}')
    search = input("SKU: ")
    for num_row in range(max_row):
        row = num_row + 1
        sku = str(sheet.cell(row=row, column=2).value).lower()
        search_sku = search.lower()
        if search_sku in sku:
            is_found = True
            onhand = int(sheet.cell(
                row=row, column=col_onhand).value) or 0
            outgoing = int(sheet.cell(
                row=row, column=col_outgoing).value) or 0
            result = onhand-outgoing
            print(f'{Fore.WHITE}{Back.GREEN}{Style.BRIGHT}สินค้าคงเหลือ: {result}{Style.RESET_ALL}')
            items = [search_sku.upper(), onhand, outgoing, result]
            export.append(items)
            check = 1
            break
        is_found = False
        if not is_found:
            execute = False
