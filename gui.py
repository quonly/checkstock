import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook
from colorama import just_fix_windows_console, init, Fore, Back, Style

just_fix_windows_console()
init()

def search_sku():
    file_path = filedialog.askopenfilename()
    file = load_workbook(file_path)
    sheet = file.worksheets[0]
    max_row = sheet.max_row
    col_onhand = int(onhand_entry.get())
    col_outgoing = int(outgoing_entry.get())
    export = []
    is_found = True
    execute = True
    search = sku_entry.get()
    while execute:
        if not is_found:
            result_label.configure(text=f'ไม่พบสินค้า', bg='red')
        for num_row in range(max_row):
            row = num_row + 1
            sku = str(sheet.cell(row=row, column=2).value).lower()
            search_sku = search.lower()
            if search_sku in sku:
                is_found = True
                onhand = int(sheet.cell(row=row, column=col_onhand).value) or 0
                outgoing = int(sheet.cell(row=row, column=col_outgoing).value) or 0
                result = onhand - outgoing
                result_label.configure(text=f'สินค้าคงเหลือ: {result}', bg='green')
                items = [search_sku.upper(), onhand, outgoing, result]
                export.append(items)
                check = 1
                break
            is_found = False
            if not is_found:
                execute = False

root = tk.Tk()
root.geometry('400x300')
root.title('Search SKU')

sku_label = tk.Label(root, text='SKU:')
sku_label.pack(pady=5)

sku_entry = tk.Entry(root)
sku_entry.pack()

onhand_label = tk.Label(root, text='คอลัมน์ onhand:')
onhand_label.pack(pady=5)

onhand_entry = tk.Entry(root)
onhand_entry.pack()

outgoing_label = tk.Label(root, text='คอลัมน์ outgoing:')
outgoing_label.pack(pady=5)

outgoing_entry = tk.Entry(root)
outgoing_entry.pack()

search_button = tk.Button(root, text='Search', command=search_sku)
search_button.pack(pady=10)

result_label = tk.Label(root, text='', font=('Arial', 16), width=20, height=2)
result_label.pack(pady=20)

root.mainloop()
